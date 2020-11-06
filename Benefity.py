import os
import openpyxl

print('Podaj nazwe 1 pliku: ')
miesiac1=input()
print('Podaj nazwe 2 pliku: ')
miesiac2=input()

    

def odczytajMiesiac(mojSlownik,plik):
    mojArkusz=openpyxl.load_workbook(plik)
    mojArkusz = mojArkusz.active
    mojSlownik = {}

    


    for wiersz in range(2,mojArkusz.max_row+1):
        if  '!' not in str(mojArkusz.cell(row=wiersz, column=3).value):
            if  mojArkusz.cell(row=wiersz, column=3).value not in mojSlownik:
                mojSlownik[(mojArkusz.cell(row=wiersz, column=3).value)] = {'kwota':[mojArkusz.cell(row=wiersz, column=7).value],'ulica':mojArkusz.cell(row=wiersz, column=11).value,'kod':mojArkusz.cell(row=wiersz, column=12).value,'miasto':mojArkusz.cell(row=wiersz, column=13).value}
            else:
                mojSlownik[(mojArkusz.cell(row=wiersz, column=3).value)]['kwota'].append(mojArkusz.cell(row=wiersz, column=7).value)

    return(mojSlownik)




maj = odczytajMiesiac("maj", miesiac1)
czerwiec = odczytajMiesiac("czerwiec", miesiac2)

wynik=openpyxl.Workbook()
ws = wynik.active

i=1
for pracownik in czerwiec:
    if pracownik in maj:
        if sum(maj[pracownik]['kwota']) != sum(czerwiec[pracownik]['kwota']):
            for k in czerwiec[pracownik]:
                #print(pracownik,k,str(adresy[pracownik]))
                ws.cell(row=i,column=1).value = pracownik
                ws.cell(row=i,column=2).value = k
                ws.cell(row=i,column=3).value = str(adresy[pracownik][0])
                ws.cell(row=i,column=4).value = str(adresy[pracownik][1])
                ws.cell(row=i,column=5).value = str(adresy[pracownik][2])
                i=i+1
    else:
        for k in czerwiec[pracownik]:
                print(pracownik,k,str(adresy[pracownik]))
                ws.cell(row=i,column=1).value = pracownik
                ws.cell(row=i,column=2).value = k
                ws.cell(row=i,column=3).value = str(adresy[pracownik][0])
                ws.cell(row=i,column=4).value = str(adresy[pracownik][1])
                ws.cell(row=i,column=5).value = str(adresy[pracownik][2])
                i=i+1


last_column=ws.max_column+1
for data in range(1,ws.max_row+1):
    ws.cell(row=data,column=last_column).value=miesiac2[8:15]
    

wynik.save('do_pythona_' + miesiac2[8:15]+'.xlsx')
print('Dane zostały zapisane do pliku do_pythona_' + miesiac2[8:15] + ' w lokalizacji: ' ,os.getcwd())
